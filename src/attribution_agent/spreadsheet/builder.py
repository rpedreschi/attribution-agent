"""Build the six-sheet board pack from a BoardPackData.

Sheets:
    1. Executive Summary     KPIs vs prior quarter, agent observations,
                             agent recommendations, top/bottom 3 by ROI
    2. Attribution by Channel  last-touch / linear / time-decay + model agreement
    3. Attribution by Campaign top 20 campaigns by attributed revenue
    4. Funnel Metrics          touch -> conversation -> MQL -> SQL -> opp -> won
    5. CAC and ROI             per program category + payback period
    6. Data & Assumptions      sources, refresh, methodology, guardrails
"""
from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from ..agent.reporting import BoardPackData

# --- palette ----------------------------------------------------------------
NAVY = "1F2A44"
BLUE = "2E5AAC"
LIGHT = "EAF0FB"
GREY = "F4F5F7"
GREEN = "1E7F4F"
RED = "B3261E"
WHITE = "FFFFFF"

TITLE_FONT = Font(name="Calibri", size=16, bold=True, color=WHITE)
H_FONT = Font(name="Calibri", size=11, bold=True, color=WHITE)
LABEL_FONT = Font(name="Calibri", size=10, bold=True, color=NAVY)
BODY_FONT = Font(name="Calibri", size=10, color="222222")
KPI_FONT = Font(name="Calibri", size=20, bold=True, color=NAVY)

TITLE_FILL = PatternFill("solid", fgColor=NAVY)
HEAD_FILL = PatternFill("solid", fgColor=BLUE)
ALT_FILL = PatternFill("solid", fgColor=GREY)
CARD_FILL = PatternFill("solid", fgColor=LIGHT)

THIN = Side(style="thin", color="D0D4DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")

MONEY = '#,##0;[Red]-#,##0'
PCT = '0.0%'
MULT = '0.00"x"'


def _title(ws: Worksheet, text: str, span: int) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    c = ws.cell(1, 1, text)
    c.font = TITLE_FONT
    c.fill = TITLE_FILL
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 30


def _header_row(ws: Worksheet, row: int, headers: list[str]) -> None:
    for col, text in enumerate(headers, start=1):
        c = ws.cell(row, col, text)
        c.font = H_FONT
        c.fill = HEAD_FILL
        c.alignment = CENTER
        c.border = BORDER
    ws.row_dimensions[row].height = 22


def _widths(ws: Worksheet, widths: list[int]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _data_row(ws: Worksheet, row: int, values: list, *, fmt: dict[int, str] | None = None,
              alt: bool = False) -> None:
    fmt = fmt or {}
    for col, val in enumerate(values, start=1):
        c = ws.cell(row, col, val)
        c.font = BODY_FONT
        c.border = BORDER
        if alt:
            c.fill = ALT_FILL
        if col in fmt:
            c.number_format = fmt[col]
            c.alignment = RIGHT
        else:
            c.alignment = LEFT


# --- sheet builders ---------------------------------------------------------

def _exec_summary(wb: Workbook, d: BoardPackData) -> None:
    ws = wb.active
    ws.title = "Executive Summary"
    _widths(ws, [26, 20, 20, 22, 30])
    _title(ws, f"{d.customer_name} — Marketing Attribution Board Pack ({d.fiscal_period})", 5)

    # KPI cards.
    kpis = [
        ("Attributed Revenue", f"${d.total_attributed:,.0f}", f"{d.revenue_qoq:+.1%} QoQ"),
        ("Blended ROI", f"{d.blended_roi:.2f}x", f"Spend ${d.total_spend:,.0f}"),
        ("Blended CAC", f"${d.blended_cac:,.0f}", f"{d.won_deals} new customers"),
        ("Model Agreement", f"{d.model_agreement:.2f}", "1.00 = models agree"),
    ]
    row = 3
    for col, (label, value, sub) in enumerate(kpis, start=1):
        lc = ws.cell(row, col, label); lc.font = LABEL_FONT; lc.fill = CARD_FILL; lc.alignment = CENTER; lc.border = BORDER
        vc = ws.cell(row + 1, col, value); vc.font = KPI_FONT; vc.fill = CARD_FILL; vc.alignment = CENTER; vc.border = BORDER
        sc = ws.cell(row + 2, col, sub); sc.font = BODY_FONT; sc.fill = CARD_FILL; sc.alignment = CENTER; sc.border = BORDER
    ws.row_dimensions[row + 1].height = 30

    # Agent observations.
    r = row + 4
    ws.cell(r, 1, "Agent Observations").font = Font(size=12, bold=True, color=NAVY)
    r += 1
    for obs in d.observations:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        c = ws.cell(r, 1, f"•  {obs}"); c.font = BODY_FONT; c.alignment = LEFT
        ws.row_dimensions[r].height = 28
        r += 1

    # Agent recommendations (pending human approval).
    r += 1
    ws.cell(r, 1, "Agent Recommendations  (pending human approval)").font = Font(size=12, bold=True, color=NAVY)
    r += 1
    _header_row(ws, r, ["Channel", "Action", "Δ Weekly Spend", "Est. Revenue Impact", "Rationale"])
    r += 1
    if d.recommendations:
        for i, rec in enumerate(d.recommendations):
            _data_row(ws, r, [rec.channel, rec.action, rec.delta, rec.expected_revenue_impact, rec.rationale],
                      fmt={3: MONEY, 4: MONEY}, alt=(i % 2 == 1))
            ws.row_dimensions[r].height = 42
            r += 1
    else:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        ws.cell(r, 1, "No material change this period — no reallocation proposed.").font = BODY_FONT
        r += 1

    # Top / bottom 3 by ROI.
    r += 1
    ranked = d.ranked_by_roi()
    ws.cell(r, 1, "Top 3 Programs by ROI").font = LABEL_FONT
    ws.cell(r, 3, "Bottom 3 Programs by ROI").font = LABEL_FONT
    r += 1
    bottom = list(reversed(ranked[-3:]))
    for i in range(3):
        t = ranked[i]
        ws.cell(r + i, 1, t.program_category).font = BODY_FONT
        c = ws.cell(r + i, 2, t.roi); c.number_format = MULT; c.font = Font(size=10, bold=True, color=GREEN); c.alignment = RIGHT
        b = bottom[i]
        ws.cell(r + i, 3, b.program_category).font = BODY_FONT
        c = ws.cell(r + i, 4, b.roi); c.number_format = MULT; c.font = Font(size=10, bold=True, color=RED); c.alignment = RIGHT
    ws.sheet_view.showGridLines = False


def _attribution_by_channel(wb: Workbook, d: BoardPackData) -> None:
    ws = wb.create_sheet("Attribution by Channel")
    _widths(ws, [22, 16, 16, 16, 12, 12, 12])
    _title(ws, "Attribution by Channel — Three Models Side by Side", 7)
    note = (f"Each model distributes the same ${d.total_attributed:,.0f} of closed-won revenue. "
            f"Model-agreement coefficient: {d.model_agreement:.2f} "
            f"(avg. pairwise correlation of channel shares).")
    ws.merge_cells("A2:G2"); c = ws.cell(2, 1, note); c.font = Font(size=9, italic=True, color="555555"); c.alignment = LEFT
    ws.row_dimensions[2].height = 26

    _header_row(ws, 3, ["Channel", "Last Touch", "Linear", "Time Decay",
                        "LT Share", "Lin Share", "TD Share"])
    tot_lt = sum(c.last_touch for c in d.channels)
    tot_ln = sum(c.linear for c in d.channels)
    tot_td = sum(c.time_decay for c in d.channels)
    row = 4
    for i, ch in enumerate(sorted(d.channels, key=lambda c: c.time_decay, reverse=True)):
        _data_row(ws, row, [ch.channel, ch.last_touch, ch.linear, ch.time_decay,
                            ch.last_touch / tot_lt, ch.linear / tot_ln, ch.time_decay / tot_td],
                  fmt={2: MONEY, 3: MONEY, 4: MONEY, 5: PCT, 6: PCT, 7: PCT}, alt=(i % 2 == 1))
        row += 1
    # totals
    _data_row(ws, row, ["Total", tot_lt, tot_ln, tot_td, 1, 1, 1],
              fmt={2: MONEY, 3: MONEY, 4: MONEY, 5: PCT, 6: PCT, 7: PCT})
    for col in range(1, 8):
        ws.cell(row, col).font = Font(size=10, bold=True, color=NAVY)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"


def _attribution_by_campaign(wb: Workbook, d: BoardPackData) -> None:
    ws = wb.create_sheet("Attribution by Campaign")
    _widths(ws, [34, 18, 16, 18, 14])
    _title(ws, "Attribution by Campaign — Top 20 by Attributed Revenue", 5)
    _header_row(ws, 2, ["Campaign", "Channel", "Spend", "Attributed Revenue", "ROI"])
    ranked = sorted(d.campaigns, key=lambda c: c.attributed_revenue, reverse=True)[:20]
    row = 3
    for i, c in enumerate(ranked):
        roi = c.roi if c.roi is not None else 0
        _data_row(ws, row, [c.name, c.channel, c.spend, c.attributed_revenue, roi],
                  fmt={3: MONEY, 4: MONEY, 5: MULT}, alt=(i % 2 == 1))
        row += 1
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"


def _funnel(wb: Workbook, d: BoardPackData) -> None:
    ws = wb.create_sheet("Funnel Metrics")
    _widths(ws, [20, 12, 14, 10, 10, 10, 8, 14, 12])
    _title(ws, "Funnel Metrics by Program Category", 9)
    _header_row(ws, 2, ["Program", "Touches", "Conversations", "MQL", "SQL", "Opp", "Won",
                        "Conv/Touch", "Won/Opp"])
    row = 3
    tot = [0] * 6
    for i, f in enumerate(sorted(d.funnel, key=lambda f: f.won, reverse=True)):
        rates = f.rates()
        _data_row(ws, row, [f.program_category, f.touches, f.conversations, f.mqls, f.sqls,
                            f.opps, f.won, rates["conv_per_touch"], rates["won_per_opp"]],
                  fmt={2: '#,##0', 3: '#,##0', 4: '#,##0', 5: '#,##0', 6: '#,##0', 7: '#,##0',
                       8: PCT, 9: PCT}, alt=(i % 2 == 1))
        for j, v in enumerate([f.touches, f.conversations, f.mqls, f.sqls, f.opps, f.won]):
            tot[j] += v
        row += 1
    _data_row(ws, row, ["Total", *tot, "", ""], fmt={c: '#,##0' for c in range(2, 8)})
    for col in range(1, 10):
        ws.cell(row, col).font = Font(size=10, bold=True, color=NAVY)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"


def _cac_roi(wb: Workbook, d: BoardPackData) -> None:
    ws = wb.create_sheet("CAC and ROI")
    _widths(ws, [22, 16, 20, 14, 16, 12, 16])
    _title(ws, "CAC and ROI by Program Category", 7)
    sub = "ROI = attributed revenue / spend (time-decay). CAC = spend / new customers. Payback = 12 / ROI months (ARR)."
    ws.merge_cells("A2:G2"); c = ws.cell(2, 1, sub); c.font = Font(size=9, italic=True, color="555555"); c.alignment = LEFT
    _header_row(ws, 3, ["Program", "Spend", "Attributed Revenue", "Deals", "CAC", "ROI", "Payback (mo)"])
    row = 4
    for i, r in enumerate(d.ranked_by_roi()):
        _data_row(ws, row, [r.program_category, r.spend, r.attributed_revenue, r.attributed_deals,
                            r.cac or 0, r.roi or 0, r.payback_months or 0],
                  fmt={2: MONEY, 3: MONEY, 4: '#,##0', 5: MONEY, 6: MULT, 7: '0.0'}, alt=(i % 2 == 1))
        row += 1
    _data_row(ws, row, ["Blended", d.total_spend, d.total_attributed, d.won_deals,
                        d.blended_cac, d.blended_roi, 12 / d.blended_roi],
              fmt={2: MONEY, 3: MONEY, 4: '#,##0', 5: MONEY, 6: MULT, 7: '0.0'})
    for col in range(1, 8):
        ws.cell(row, col).font = Font(size=10, bold=True, color=NAVY)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"


def _data_assumptions(wb: Workbook, d: BoardPackData) -> None:
    ws = wb.create_sheet("Data & Assumptions")
    _widths(ws, [30, 70])
    _title(ws, "Data & Assumptions", 2)
    rows = [
        ("Customer", d.customer_name),
        ("Reporting period", d.fiscal_period),
        ("Generated", date.today().isoformat()),
        ("", ""),
        ("Source systems", "Salesforce CRM, HubSpot MAP, Google Analytics 4, "
                           "LinkedIn Ads, Google Ads, Outreach"),
        ("Pipeline", "Source APIs -> Kafka (Confluent Cloud) -> DeltaStream "
                     "(identity resolution + materialized views) -> MCP -> Agent -> this workbook"),
        ("Refresh cadence", "DeltaStream materialized views update continuously; board pack generated weekly"),
        ("Identity resolution", "Anonymous web user_id stitched to Salesforce contact/account via "
                                "the HubSpot form web-cookie bridge and Salesforce contacts CDC"),
        ("", ""),
        ("Attribution models", "Last touch (100% to final touch), Linear (equal credit), "
                               "Time decay (7-day half-life). Each distributes the same closed-won revenue."),
        ("CAC", "Spend / new customers (attributed deals)"),
        ("ROI", "Attributed revenue / spend, using the time-decay model"),
        ("Payback", "12 / ROI, treating attributed revenue as ARR"),
        ("Model agreement", f"{d.model_agreement:.2f} — avg. pairwise correlation of channel shares across the three models"),
        ("", ""),
        ("Agent guardrails", "Agent recommends, human approves — no autonomous spend changes in v1."),
        ("  Reallocation cap", "±20% of current channel spend per week"),
        ("  Thin-channel rule", "Channels with <30 conversions in trailing 90 days are excluded"),
        ("  Excluded channels", "Brand and Events are excluded from agent autonomy entirely"),
        ("  Audit", "Every recommendation is logged with rationale and source-data references"),
    ]
    r = 2
    for label, val in rows:
        lc = ws.cell(r, 1, label); lc.font = LABEL_FONT; lc.alignment = LEFT
        vc = ws.cell(r, 2, val); vc.font = BODY_FONT; vc.alignment = LEFT
        if val and len(val) > 60:
            ws.row_dimensions[r].height = 30
        r += 1
    ws.sheet_view.showGridLines = False


def build_workbook(data: BoardPackData, out_path: Path) -> Path:
    wb = Workbook()
    _exec_summary(wb, data)
    _attribution_by_channel(wb, data)
    _attribution_by_campaign(wb, data)
    _funnel(wb, data)
    _cac_roi(wb, data)
    _data_assumptions(wb, data)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path
