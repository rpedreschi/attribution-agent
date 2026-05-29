"""End-to-end checks on the sample pipeline and the workbook artifact."""
from pathlib import Path

from openpyxl import load_workbook

from attribution_agent.agent.recommendations import RecommendationEngine
from attribution_agent.agent.guardrails import Guardrails
from attribution_agent.agent.reporting import BoardPackData
from attribution_agent.agent.runtime import run_pipeline
from attribution_agent.config import GuardrailConfig
from attribution_agent.spreadsheet.builder import build_workbook


def test_board_pack_data_ties_out():
    d = BoardPackData.from_sample()
    assert d.total_attributed == 4_280_000
    for total in (sum(c.last_touch for c in d.channels),
                  sum(c.linear for c in d.channels),
                  sum(c.time_decay for c in d.channels)):
        assert total == 4_280_000
    assert round(d.blended_roi, 2) == 2.25
    assert 0.0 <= d.model_agreement <= 1.0


def test_recommendations_respect_guardrails():
    d = BoardPackData.from_sample()
    cfg = GuardrailConfig()
    recs = RecommendationEngine(Guardrails(cfg)).propose(d)
    assert recs, "expected a material reallocation on the sample data"
    for r in recs:
        # No excluded channel ever appears.
        assert r.channel not in cfg.excluded_channels
        # Every move is within the ±20% weekly cap.
        assert abs(r.delta) <= r.current_spend * cfg.max_weekly_reallocation_pct + 1e-6
        assert r.rationale and r.source_refs


def test_pipeline_and_workbook(tmp_path: Path):
    data = run_pipeline(source="sample")
    assert data.observations
    out = build_workbook(data, tmp_path / "pack.xlsx")
    wb = load_workbook(out)
    assert wb.sheetnames == [
        "Executive Summary", "Attribution by Channel", "Attribution by Campaign",
        "Funnel Metrics", "CAC and ROI", "Data & Assumptions",
    ]
    # Channel sheet total row equals $4.28M under each model.
    ws = wb["Attribution by Channel"]
    assert [ws.cell(ws.max_row, c).value for c in range(2, 5)] == [4_280_000] * 3
